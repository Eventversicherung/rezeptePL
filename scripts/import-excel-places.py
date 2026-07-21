#!/usr/bin/env python3
"""Import feste Polenläden from Excel → SQL seed for public.places.

Geocoding: Photon (Komoot / OSM), Nominatim fallback.
Usage:
  python3 scripts/import-excel-places.py \\
    [--xlsx data/polenlaeden_feste_standorte_deutschland.xlsx] \\
    > supabase/seeds/places_curated_xlsx.sql
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

UA = "AlemniamPlaceImport/1.0 (curated Excel; https://github.com/Eventversicherung/rezeptePL)"
PHOTON = "https://photon.komoot.io/api/"
NOMINATIM = "https://nominatim.openstreetmap.org/search"


def esc(s: str | None) -> str:
    if s is None:
        return "null"
    return "'" + str(s).replace("'", "''") + "'"


def slugify(s: str) -> str:
    s = s.lower()
    repl = {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
        "ł": "l",
        "ń": "n",
        "ś": "s",
        "ź": "z",
        "ż": "z",
        "ć": "c",
        "ę": "e",
        "ą": "a",
    }
    for a, b in repl.items():
        s = s.replace(a, b)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:48] or "place"


def http_json(url: str, timeout: int = 30) -> dict | list:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode())


def geocode_photon(street: str, postal: str, city: str) -> tuple[float, float] | None:
    q = f"{street}, {postal} {city}, Germany"
    url = (
        f"{PHOTON}?q={urllib.parse.quote(q)}"
        f"&bbox=5.8,47.2,15.1,55.1&limit=5&lang=de"
    )
    try:
        data = http_json(url)
    except Exception as exc:  # noqa: BLE001
        print(f"-- photon fail: {exc}", file=sys.stderr)
        return None
    features = data.get("features") if isinstance(data, dict) else []
    best = None
    for f in features or []:
        props = f.get("properties") or {}
        coords = (f.get("geometry") or {}).get("coordinates") or []
        if len(coords) < 2:
            continue
        lon, lat = float(coords[0]), float(coords[1])
        cc = (props.get("countrycode") or "").upper()
        if cc and cc not in ("DE", "DEU"):
            continue
        post = str(props.get("postcode") or "")
        score = 0
        if post and post == str(postal):
            score += 3
        city_l = city.lower()
        for key in ("city", "town", "village", "name"):
            v = (props.get(key) or "").lower()
            if v and (city_l in v or v in city_l):
                score += 2
                break
        if props.get("street"):
            score += 1
        if best is None or score > best[0]:
            best = (score, lat, lon)
    if best and best[0] >= 2:
        return best[1], best[2]
    if best:
        return best[1], best[2]
    return None


def geocode_nominatim(street: str, postal: str, city: str) -> tuple[float, float] | None:
    params = urllib.parse.urlencode(
        {
            "street": street,
            "postalcode": postal,
            "city": city,
            "country": "Germany",
            "format": "json",
            "limit": 1,
        }
    )
    try:
        data = http_json(f"{NOMINATIM}?{params}")
    except Exception as exc:  # noqa: BLE001
        print(f"-- nominatim fail: {exc}", file=sys.stderr)
        return None
    if not isinstance(data, list) or not data:
        return None
    return float(data[0]["lat"]), float(data[0]["lon"])


def load_rows(xlsx: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Standorte"]
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    for r in raw[2:]:
        if r[0] is None:
            continue
        out.append(
            {
                "id": int(r[0]),
                "category": (r[1] or "").strip(),
                "name": (r[2] or "").strip(),
                "street": (r[3] or "").strip(),
                "postal": str(r[4] or "").strip(),
                "city": (r[5] or "").strip(),
                "state": (r[6] or "").strip(),
                "hours": (r[7] or "").strip(),
                "website": (r[8] or None),
                "notes": (r[9] or "").strip() if r[9] else "",
                "since": (r[10] or "").strip() if r[10] else "",
                "assortment": (r[11] or "").strip() if r[11] else "",
                "source": (r[12] or "").strip() if r[12] else "",
                "source_status": (r[13] or "").strip() if r[13] else "",
            }
        )
    return out


def build_description(row: dict) -> tuple[str, str]:
    parts_de = []
    if row["notes"]:
        parts_de.append(row["notes"])
    if row["assortment"]:
        parts_de.append(row["assortment"])
    if row["since"]:
        parts_de.append(f"Seit {row['since']}.")
    de = " ".join(parts_de).strip() or "Fester Polenladen / polnischer Fachmarkt."
    # Keep PL short; UI is bilingual but names stay original.
    pl = de
    return de[:800], pl[:800]


def pick_source_url(row: dict) -> str | None:
    website = row["website"]
    source = row["source"] or ""
    if website and str(website).startswith("http"):
        return str(website)
    if source.startswith("http") and "google.com/search" not in source:
        return source
    return website if website else (source or None)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default="data/polenlaeden_feste_standorte_deutschland.xlsx",
    )
    ap.add_argument("--cache", default="/tmp/alemniam_geocode_cache.json")
    args = ap.parse_args()
    xlsx = Path(args.xlsx)
    if not xlsx.is_absolute():
        xlsx = Path(__file__).resolve().parents[1] / xlsx

    cache_path = Path(args.cache)
    cache: dict = {}
    if cache_path.exists():
        cache = json.loads(cache_path.read_text())

    rows = load_rows(xlsx)
    print(f"-- excel rows: {len(rows)}", file=sys.stderr)

    geocoded = []
    failed = []
    for i, row in enumerate(rows):
        key = f"{row['street']}|{row['postal']}|{row['city']}"
        if key in cache:
            lat, lng = cache[key]
        else:
            hit = geocode_photon(row["street"], row["postal"], row["city"])
            time.sleep(0.35)
            if not hit:
                hit = geocode_nominatim(row["street"], row["postal"], row["city"])
                time.sleep(1.1)
            if not hit:
                failed.append(row)
                print(f"-- FAIL geocode #{row['id']} {row['name']} {key}", file=sys.stderr)
                continue
            lat, lng = hit
            cache[key] = [lat, lng]
            if (i + 1) % 10 == 0:
                cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2))
                print(f"-- geocode {i+1}/{len(rows)}", file=sys.stderr)
        row = {**row, "lat": lat, "lng": lng}
        geocoded.append(row)

    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2))
    print(f"-- geocoded {len(geocoded)} failed {len(failed)}", file=sys.stderr)

    print("-- Curated Excel import: feste Polenläden DE")
    print("-- Source file: data/polenlaeden_feste_standorte_deutschland.xlsx")
    print("begin;")
    print("delete from public.places where source = 'curated';")
    print(
        "create temporary table tmp_curated ("
        "slug text primary key, kind public.place_kind, lat float8, lng float8, "
        "street text, postal_code text, city text, state text, website text, "
        "opening_hours jsonb, tags text[], name_de text, name_pl text, "
        "desc_de text, desc_pl text, source_url text);"
    )

    vals = []
    for row in geocoded:
        slug = f"curated-xlsx-{row['id']}-{slugify(row['name'])}"[:80]
        website = row["website"]
        if website and not str(website).startswith("http"):
            website = "https://" + str(website)
        hours = row["hours"] or ""
        oh = json.dumps({"raw": hours} if hours else {}, ensure_ascii=False)
        desc_de, desc_pl = build_description(row)
        tags = ["xlsx", "feste-standorte"]
        if "markt" in (row["category"] or "").lower() and "fachmarkt" not in (
            row["category"] or ""
        ).lower():
            # still shops (feste Geschäfte), not open-air markets
            tags.append("fachmarkt")
        source_url = pick_source_url(row)
        vals.append(
            "("
            f"{esc(slug)}, 'shop'::public.place_kind, {row['lat']}, {row['lng']}, "
            f"{esc(row['street'])}, {esc(row['postal'])}, {esc(row['city'])}, "
            f"{esc(row['state'])}, {esc(website) if website else 'null'}, "
            f"{esc(oh)}::jsonb, array[{','.join(esc(t) for t in tags)}]::text[], "
            f"{esc(row['name'])}, {esc(row['name'])}, "
            f"{esc(desc_de)}, {esc(desc_pl)}, "
            f"{esc(source_url) if source_url else 'null'}"
            ")"
        )

    print("insert into tmp_curated values\n" + ",\n".join(vals) + ";")
    print(
        """
insert into public.places (
  slug, kind, status, lat, lng, street, postal_code, city, state,
  website, phone, opening_hours, tags, source, source_url, verified_at
)
select slug, kind, 'published', lat, lng, street, postal_code, city, state,
  website, null, opening_hours, tags, 'curated'::public.place_source, source_url, now()
from tmp_curated
on conflict (slug) do update set
  kind = excluded.kind, status = 'published', lat = excluded.lat, lng = excluded.lng,
  street = excluded.street, postal_code = excluded.postal_code, city = excluded.city,
  state = excluded.state, website = excluded.website,
  opening_hours = excluded.opening_hours, tags = excluded.tags, source = excluded.source,
  source_url = excluded.source_url, verified_at = now(), updated_at = now();

insert into public.place_translations (place_id, locale, name, description, schedule_note, seo_title, seo_description)
select p.id, 'de', s.name_de, s.desc_de, '', s.name_de || ' | Alemniam', left(s.desc_de, 155)
from tmp_curated s join public.places p on p.slug = s.slug
on conflict (place_id, locale) do update set
  name = excluded.name, description = excluded.description,
  seo_title = excluded.seo_title, seo_description = excluded.seo_description;

insert into public.place_translations (place_id, locale, name, description, schedule_note, seo_title, seo_description)
select p.id, 'pl', s.name_pl, s.desc_pl, '', s.name_pl || ' | Alemniam', left(s.desc_pl, 155)
from tmp_curated s join public.places p on p.slug = s.slug
on conflict (place_id, locale) do update set
  name = excluded.name, description = excluded.description,
  seo_title = excluded.seo_title, seo_description = excluded.seo_description;

-- Drop OSM duplicates within ~120m of a curated shop (Excel is richer).
delete from public.places osm
where osm.source = 'osm'
  and exists (
    select 1 from public.places c
    where c.source = 'curated'
      and st_dwithin(osm.location, c.location, 120)
  );

commit;
"""
    )


if __name__ == "__main__":
    main()
